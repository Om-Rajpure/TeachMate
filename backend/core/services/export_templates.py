"""
TeachMate — Export Template Service
=====================================
Generates beautifully formatted Excel mark-entry templates for teachers.

Functions:
    generate_theory_template(subject, division, students)   -> openpyxl.Workbook
    generate_practical_template(subject, division, batch, students, experiments) -> openpyxl.Workbook
"""

import datetime
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


# ────────────────────────────────────────────────────────────
#  Shared colour palette
# ────────────────────────────────────────────────────────────
THEORY_TITLE_FILL   = PatternFill("solid", fgColor="1A56DB")   # rich blue
THEORY_HEADER_FILL  = PatternFill("solid", fgColor="DBEAFE")   # light sky
THEORY_META_FILL    = PatternFill("solid", fgColor="EFF6FF")   # very light blue
THEORY_ALT_FILL     = PatternFill("solid", fgColor="F8FAFC")   # near-white alt row

PRAC_TITLE_FILL     = PatternFill("solid", fgColor="065F46")   # deep green
PRAC_HEADER_FILL    = PatternFill("solid", fgColor="D1FAE5")   # mint
PRAC_META_FILL      = PatternFill("solid", fgColor="ECFDF5")   # very light green
PRAC_EXP_A_FILL     = PatternFill("solid", fgColor="DBEAFE")   # blue-50
PRAC_EXP_B_FILL     = PatternFill("solid", fgColor="EDE9FE")   # purple-50
PRAC_ASSIGN_FILL    = PatternFill("solid", fgColor="D1FAE5")   # green-50
PRAC_TOTAL_FILL     = PatternFill("solid", fgColor="FEF9C3")   # yellow-50
PRAC_ALT_FILL       = PatternFill("solid", fgColor="F8FAFC")

WHITE_FONT  = Font(color="FFFFFF", bold=True, name="Calibri")
DARK_FONT   = Font(color="1E293B", bold=True, name="Calibri")
MUTED_FONT  = Font(color="64748B", name="Calibri")
META_FONT   = Font(color="1E3A5F", bold=True, name="Calibri")
GREEN_FONT  = Font(color="065F46", bold=True, name="Calibri")

TODAY_STR   = datetime.date.today().strftime("%d %B %Y")


def _thin_border(top=True, bottom=True, left=True, right=True):
    side = Side(style="thin", color="CBD5E1")
    return Border(
        top    = side if top    else Side(),
        bottom = side if bottom else Side(),
        left   = side if left   else Side(),
        right  = side if right  else Side(),
    )


def _center(wrap=False):
    return Alignment(horizontal="center", vertical="center", wrap_text=wrap)


def _left():
    return Alignment(horizontal="left", vertical="center")


def _set_col_width(ws, col_idx, width):
    ws.column_dimensions[get_column_letter(col_idx)].width = width


# ════════════════════════════════════════════════════════════
#  THEORY TEMPLATE
# ════════════════════════════════════════════════════════════

def generate_theory_template(subject, division, students):
    """
    Returns an openpyxl Workbook with two sheets:
      1. 'Marks Entry'   — pre-filled student list + IA-1 / IA-2 columns
      2. 'Instructions'  — usage notes
    """
    wb = Workbook()

    # ── Sheet 1: Marks Entry ──────────────────────────────────
    ws = wb.active
    ws.title = "Marks Entry"
    ws.sheet_view.showGridLines = False

    # ── Row 1: Grand Title (merged A1:E1) ────────────────────
    ws.merge_cells("A1:E1")
    title_cell = ws["A1"]
    title_cell.value   = "TeachMate — Theory Marks Entry Template"
    title_cell.fill    = THEORY_TITLE_FILL
    title_cell.font    = Font(color="FFFFFF", bold=True, size=14, name="Calibri")
    title_cell.alignment = _center()
    ws.row_dimensions[1].height = 36

    # ── Row 2: Meta labels ───────────────────────────────────
    meta_labels = ["Subject", "Division", "Type", "Generated On"]
    meta_values = [
        f"{subject.name} ({subject.code})",
        division.name,
        "Theory",
        TODAY_STR,
    ]
    for col_offset, (label, val) in enumerate(zip(meta_labels, meta_values)):
        lc = ws.cell(row=2, column=col_offset + 1)
        lc.value     = label
        lc.fill      = THEORY_META_FILL
        lc.font      = MUTED_FONT
        lc.alignment = _center()
        lc.border    = _thin_border()

        vc = ws.cell(row=3, column=col_offset + 1)
        vc.value     = val
        vc.fill      = THEORY_META_FILL
        vc.font      = META_FONT
        vc.alignment = _center()
        vc.border    = _thin_border()

    ws.row_dimensions[2].height = 20
    ws.row_dimensions[3].height = 20

    # ── Row 4: Empty spacer ──────────────────────────────────
    ws.row_dimensions[4].height = 10

    # ── Row 5: Column Headers ────────────────────────────────
    HEADER_ROW = 5
    headers = ["Roll No", "Student Name", "IA-1\n(Out of 20)", "IA-2\n(Out of 20)", "Average\n(Auto)"]
    header_widths = [10, 30, 16, 16, 14]

    for col, (h, w) in enumerate(zip(headers, header_widths), start=1):
        cell = ws.cell(row=HEADER_ROW, column=col)
        cell.value     = h
        cell.fill      = THEORY_HEADER_FILL
        cell.font      = Font(color="1E40AF", bold=True, name="Calibri", size=10)
        cell.alignment = _center(wrap=True)
        cell.border    = _thin_border()
        _set_col_width(ws, col, w)

    ws.row_dimensions[HEADER_ROW].height = 32

    # ── Student rows ─────────────────────────────────────────
    ia1_cells = []
    ia2_cells = []

    for idx, student in enumerate(students):
        row = HEADER_ROW + 1 + idx
        link = student.subject_links.filter(subject=subject, division=division).first()
        roll = link.roll_number if link else (idx + 1)

        bg = THEORY_ALT_FILL if idx % 2 == 1 else PatternFill("solid", fgColor="FFFFFF")

        roll_cell = ws.cell(row=row, column=1, value=roll)
        roll_cell.fill = bg; roll_cell.font = Font(bold=True, color="1E40AF", name="Calibri")
        roll_cell.alignment = _center(); roll_cell.border = _thin_border()

        name_cell = ws.cell(row=row, column=2, value=student.name)
        name_cell.fill = bg; name_cell.font = Font(name="Calibri")
        name_cell.alignment = _left(); name_cell.border = _thin_border()

        ia1 = ws.cell(row=row, column=3)
        ia1.fill = bg; ia1.alignment = _center(); ia1.border = _thin_border()
        ia1_cells.append(ia1)

        ia2 = ws.cell(row=row, column=4)
        ia2.fill = bg; ia2.alignment = _center(); ia2.border = _thin_border()
        ia2_cells.append(ia2)

        # Average formula (auto-compute)
        avg_cell = ws.cell(row=row, column=5)
        avg_cell.value     = f"=IF(AND(C{row}=\"\",D{row}=\"\"),\"\",IFERROR((IF(C{row}=\"\",0,C{row})+IF(D{row}=\"\",0,D{row}))/2,\"\"))"
        avg_cell.fill      = PatternFill("solid", fgColor="EFF6FF")
        avg_cell.font      = Font(color="1E40AF", bold=True, name="Calibri")
        avg_cell.alignment = _center()
        avg_cell.border    = _thin_border()
        avg_cell.number_format = "0.0"

        ws.row_dimensions[row].height = 22

    data_start = HEADER_ROW + 1
    data_end   = HEADER_ROW + len(students) if students else HEADER_ROW + 1

    # ── Data Validation: IA columns (0–20) ──────────────────
    dv_ia = DataValidation(
        type="decimal",
        operator="between",
        formula1=0,
        formula2=20,
        allow_blank=True,
        showErrorMessage=True,
        errorTitle="Invalid Marks",
        error="Please enter a number between 0 and 20.",
        errorStyle="warning",
    )
    dv_ia.sqref = f"C{data_start}:D{data_end}"
    ws.add_data_validation(dv_ia)

    # ── Freeze panes below header, right of name ─────────────
    ws.freeze_panes = f"C{HEADER_ROW + 1}"

    # ── Sheet 2: Instructions ─────────────────────────────────
    wi = wb.create_sheet(title="Instructions")
    wi.sheet_view.showGridLines = False

    wi.merge_cells("A1:C1")
    t = wi["A1"]
    t.value = "📋  How to use this template"
    t.fill  = THEORY_TITLE_FILL
    t.font  = Font(color="FFFFFF", bold=True, size=13, name="Calibri")
    t.alignment = _center()
    wi.row_dimensions[1].height = 32

    instructions = [
        ("1", "Fill only the IA-1 and IA-2 columns. Leave blank if marks not yet available."),
        ("2", "Do NOT modify the Roll No or Student Name columns."),
        ("3", "Do NOT rename or delete any column headers."),
        ("4", "Marks must be between 0 and 20. Entering more will show a warning."),
        ("5", "The 'Average' column is auto-calculated — do not type in it."),
        ("6", "Save the file as .xlsx format before uploading."),
        ("7", "Upload using the Bulk Upload button on the Marks Management page."),
    ]

    for i, (num, text) in enumerate(instructions, start=2):
        wi.row_dimensions[i].height = 22
        num_cell = wi.cell(row=i, column=1, value=num)
        num_cell.font = Font(bold=True, color="1A56DB", name="Calibri")
        num_cell.alignment = _center()
        num_cell.fill = PatternFill("solid", fgColor="EFF6FF")
        num_cell.border = _thin_border()

        text_cell = wi.cell(row=i, column=2, value=text)
        text_cell.font = Font(name="Calibri")
        text_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        text_cell.border = _thin_border()
        wi.merge_cells(f"B{i}:C{i}")

    wi.column_dimensions["A"].width = 5
    wi.column_dimensions["B"].width = 70
    wi.column_dimensions["C"].width = 10

    return wb


# ════════════════════════════════════════════════════════════
#  PRACTICAL TEMPLATE
# ════════════════════════════════════════════════════════════

def generate_practical_template(subject, division, batch, students, experiments):
    """
    Returns an openpyxl Workbook with two sheets:
      1. 'Practical Marks' — experiment + assignment columns, total formulas
      2. 'Instructions'    — usage notes

    experiments: list of Experiment model instances (ordered by experiment_number)
    """
    wb = Workbook()

    # ── Sheet 1: Practical Marks ──────────────────────────────
    ws = wb.active
    ws.title = "Practical Marks"
    ws.sheet_view.showGridLines = False

    # ── Build header column list dynamically ─────────────────
    # Fixed columns: Roll No (1), Student Name (2)
    # Per experiment: Exp-N A, Exp-N B, Exp-N C, Exp-N D, Exp-N Total
    # Assignments: Ass1 A, Ass1 B, Ass1 C, Ass1 Total, Ass2 A, Ass2 B, Ass2 C, Ass2 Total
    # Grand total col at end (Exp Avg | Assign Avg)

    EXP_PARTS   = [("A", 3), ("B", 4), ("C", 4), ("D", 4)]  # (label, max)
    ASS_PARTS   = [("A", 2), ("B", 2), ("C", 1)]             # (label, max)

    FIXED_COLS  = 2
    EXP_STRIDE  = len(EXP_PARTS) + 1          # 4 parts + 1 total
    ASS_STRIDE  = len(ASS_PARTS) + 1           # 3 parts + 1 total
    NUM_EXPS    = len(experiments)
    HAS_ASSIGNS = subject.has_assignments

    EXP_COLS_START   = FIXED_COLS + 1
    ASS_COLS_START   = EXP_COLS_START + NUM_EXPS * EXP_STRIDE
    EXP_AVG_COL      = ASS_COLS_START - 1  # last col of experiments block? No:
    # Corrected layout: after all exp blocks, add Exp Avg col, then assignments
    EXP_AVG_COL      = EXP_COLS_START + NUM_EXPS * EXP_STRIDE
    ASS_COLS_START   = EXP_AVG_COL + 1 if HAS_ASSIGNS else None
    ASSIGN_AVG_COL   = (ASS_COLS_START + 2 * ASS_STRIDE) if HAS_ASSIGNS else None

    TOTAL_COLS       = EXP_AVG_COL + (2 * ASS_STRIDE + 1 if HAS_ASSIGNS else 0)

    # ── Row 1: Grand Title ────────────────────────────────────
    title_end_col = max(TOTAL_COLS, 6)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=title_end_col)
    title_cell = ws["A1"]
    title_cell.value     = "TeachMate — Practical Marks Entry Template"
    title_cell.fill      = PRAC_TITLE_FILL
    title_cell.font      = Font(color="FFFFFF", bold=True, size=14, name="Calibri")
    title_cell.alignment = _center()
    ws.row_dimensions[1].height = 36

    # ── Row 2–3: Metadata ─────────────────────────────────────
    meta_labels = ["Subject", "Division", "Batch", "Type", "Generated On"]
    meta_values = [
        f"{subject.name} ({subject.code})",
        division.name,
        batch.name if batch else "All Batches",
        "Practical",
        TODAY_STR,
    ]
    for col_offset, (label, val) in enumerate(zip(meta_labels, meta_values)):
        lc = ws.cell(row=2, column=col_offset + 1)
        lc.value = label; lc.fill = PRAC_META_FILL
        lc.font = MUTED_FONT; lc.alignment = _center(); lc.border = _thin_border()

        vc = ws.cell(row=3, column=col_offset + 1)
        vc.value = val; vc.fill = PRAC_META_FILL
        vc.font = Font(color="065F46", bold=True, name="Calibri")
        vc.alignment = _center(); vc.border = _thin_border()

    ws.row_dimensions[2].height = 20
    ws.row_dimensions[3].height = 20

    # ── Row 4: Empty spacer ──────────────────────────────────
    ws.row_dimensions[4].height = 8

    # ── Row 5: Group label row (experiment names) ─────────────
    GROUP_ROW  = 5
    SUBHDR_ROW = 6
    DATA_START = 7

    # Roll / Name cells
    ws.merge_cells(start_row=GROUP_ROW, start_column=1, end_row=SUBHDR_ROW, end_column=1)
    c = ws.cell(row=GROUP_ROW, column=1, value="Roll No")
    c.fill = PatternFill("solid", fgColor="1E3A5F"); c.font = WHITE_FONT
    c.alignment = _center(); c.border = _thin_border()

    ws.merge_cells(start_row=GROUP_ROW, start_column=2, end_row=SUBHDR_ROW, end_column=2)
    c = ws.cell(row=GROUP_ROW, column=2, value="Student Name")
    c.fill = PatternFill("solid", fgColor="1E3A5F"); c.font = WHITE_FONT
    c.alignment = _center(); c.border = _thin_border()

    # Experiment group headers
    exp_fills = [PRAC_EXP_A_FILL, PRAC_EXP_B_FILL]  # alternate per experiment
    exp_fonts = [Font(color="1E40AF", bold=True, name="Calibri"),
                 Font(color="4C1D95", bold=True, name="Calibri")]

    for exp_idx, exp in enumerate(experiments):
        col_start = EXP_COLS_START + exp_idx * EXP_STRIDE
        col_end   = col_start + len(EXP_PARTS) - 1
        fill      = exp_fills[exp_idx % 2]
        font      = exp_fonts[exp_idx % 2]

        # Merge group label
        ws.merge_cells(start_row=GROUP_ROW, start_column=col_start, end_row=GROUP_ROW, end_column=col_end + 1)
        gc = ws.cell(row=GROUP_ROW, column=col_start)
        gc.value     = f"Exp-{exp.experiment_number}"
        gc.fill      = fill
        gc.font      = font
        gc.alignment = _center()
        gc.border    = _thin_border()

        # Sub-header: A, B, C, D, Total
        for part_idx, (part_label, part_max) in enumerate(EXP_PARTS):
            sc = ws.cell(row=SUBHDR_ROW, column=col_start + part_idx)
            sc.value     = f"{part_label}\n(/{part_max})"
            sc.fill      = fill
            sc.font      = Font(color=font.color, bold=True, name="Calibri", size=9)
            sc.alignment = _center(wrap=True)
            sc.border    = _thin_border()

        total_sc = ws.cell(row=SUBHDR_ROW, column=col_start + len(EXP_PARTS))
        total_sc.value     = "Total\n(/15)"
        total_sc.fill      = PRAC_TOTAL_FILL
        total_sc.font      = Font(color="92400E", bold=True, name="Calibri", size=9)
        total_sc.alignment = _center(wrap=True)
        total_sc.border    = _thin_border()

    # Exp Avg column header
    ws.merge_cells(start_row=GROUP_ROW, start_column=EXP_AVG_COL, end_row=SUBHDR_ROW, end_column=EXP_AVG_COL)
    ea = ws.cell(row=GROUP_ROW, column=EXP_AVG_COL, value="Exp\nAvg")
    ea.fill = PatternFill("solid", fgColor="1E40AF"); ea.font = WHITE_FONT
    ea.alignment = _center(wrap=True); ea.border = _thin_border()

    # Assignment headers
    if HAS_ASSIGNS:
        for ass_idx in range(2):
            col_start = ASS_COLS_START + ass_idx * ASS_STRIDE
            col_end   = col_start + len(ASS_PARTS) - 1

            ws.merge_cells(start_row=GROUP_ROW, start_column=col_start,
                           end_row=GROUP_ROW, end_column=col_end + 1)
            gc = ws.cell(row=GROUP_ROW, column=col_start)
            gc.value     = f"Assignment {ass_idx + 1}"
            gc.fill      = PRAC_ASSIGN_FILL
            gc.font      = Font(color="065F46", bold=True, name="Calibri")
            gc.alignment = _center()
            gc.border    = _thin_border()

            for part_idx, (part_label, part_max) in enumerate(ASS_PARTS):
                sc = ws.cell(row=SUBHDR_ROW, column=col_start + part_idx)
                sc.value     = f"{part_label}\n(/{part_max})"
                sc.fill      = PRAC_ASSIGN_FILL
                sc.font      = Font(color="065F46", bold=True, name="Calibri", size=9)
                sc.alignment = _center(wrap=True)
                sc.border    = _thin_border()

            total_sc = ws.cell(row=SUBHDR_ROW, column=col_start + len(ASS_PARTS))
            total_sc.value     = "Total\n(/5)"
            total_sc.fill      = PRAC_TOTAL_FILL
            total_sc.font      = Font(color="92400E", bold=True, name="Calibri", size=9)
            total_sc.alignment = _center(wrap=True)
            total_sc.border    = _thin_border()

        # Assign Avg
        ws.merge_cells(start_row=GROUP_ROW, start_column=ASSIGN_AVG_COL,
                       end_row=SUBHDR_ROW, end_column=ASSIGN_AVG_COL)
        aa = ws.cell(row=GROUP_ROW, column=ASSIGN_AVG_COL, value="Assign\nAvg")
        aa.fill = PatternFill("solid", fgColor="065F46"); aa.font = WHITE_FONT
        aa.alignment = _center(wrap=True); aa.border = _thin_border()

    ws.row_dimensions[GROUP_ROW].height = 24
    ws.row_dimensions[SUBHDR_ROW].height = 32

    # ── Data Validations for each column group ────────────────
    # We'll add one DV per part type so messages are clear

    def _add_dv(ws, col_letter_range, max_val):
        dv = DataValidation(
            type="decimal", operator="between",
            formula1=0, formula2=max_val,
            allow_blank=True, showErrorMessage=True,
            errorTitle="Invalid Marks",
            error=f"Please enter a number between 0 and {max_val}.",
            errorStyle="warning"
        )
        dv.sqref = col_letter_range
        ws.add_data_validation(dv)

    data_end_row = DATA_START + len(students) - 1 if students else DATA_START

    # ── Student Data Rows ──────────────────────────────────────
    for s_idx, student in enumerate(students):
        row = DATA_START + s_idx
        link = student.subject_links.filter(subject=subject, division=division).first()
        if not link and batch:
            link = student.subject_links.filter(subject=subject).first()
        roll = link.roll_number if link else (s_idx + 1)

        bg = PRAC_ALT_FILL if s_idx % 2 == 1 else PatternFill("solid", fgColor="FFFFFF")

        # Roll
        rc = ws.cell(row=row, column=1, value=roll)
        rc.fill = bg; rc.font = Font(bold=True, color="1E40AF", name="Calibri")
        rc.alignment = _center(); rc.border = _thin_border()

        # Name
        nc = ws.cell(row=row, column=2, value=student.name)
        nc.fill = bg; nc.font = Font(name="Calibri")
        nc.alignment = _left(); nc.border = _thin_border()

        # Experiment cells
        exp_total_cols = []  # track total column letters for Exp Avg formula
        for exp_idx in range(NUM_EXPS):
            col_start = EXP_COLS_START + exp_idx * EXP_STRIDE
            part_col_letters = []
            for part_idx in range(len(EXP_PARTS)):
                col   = col_start + part_idx
                cell  = ws.cell(row=row, column=col)
                cell.fill = bg; cell.alignment = _center(); cell.border = _thin_border()
                part_col_letters.append(get_column_letter(col))

            # Exp total formula: =SUM(A_col:D_col for this row)
            total_col = col_start + len(EXP_PARTS)
            total_ref = f"{part_col_letters[0]}{row}:{part_col_letters[-1]}{row}"
            total_cell = ws.cell(row=row, column=total_col)
            total_cell.value       = f"=SUM({total_ref})"
            total_cell.fill        = PatternFill("solid", fgColor="FFFBEB")
            total_cell.font        = Font(color="92400E", bold=True, name="Calibri")
            total_cell.alignment   = _center()
            total_cell.border      = _thin_border()
            total_cell.number_format = "0"
            exp_total_cols.append(get_column_letter(total_col))

        # Exp Avg formula: average of all experiment totals
        exp_avg_formula = "+".join([f"{c}{row}" for c in exp_total_cols])
        exp_avg_cell = ws.cell(row=row, column=EXP_AVG_COL)
        exp_avg_cell.value = (
            f"=IFERROR(({exp_avg_formula})/{NUM_EXPS},\"\")"
            if NUM_EXPS > 0 else '""'
        )
        exp_avg_cell.fill = PatternFill("solid", fgColor="DBEAFE")
        exp_avg_cell.font = Font(color="1E40AF", bold=True, name="Calibri")
        exp_avg_cell.alignment = _center()
        exp_avg_cell.border = _thin_border()
        exp_avg_cell.number_format = "0.0"

        # Assignment cells
        if HAS_ASSIGNS:
            ass_total_cols = []
            for ass_idx in range(2):
                col_start = ASS_COLS_START + ass_idx * ASS_STRIDE
                part_col_letters = []
                for part_idx in range(len(ASS_PARTS)):
                    col  = col_start + part_idx
                    cell = ws.cell(row=row, column=col)
                    cell.fill = bg; cell.alignment = _center(); cell.border = _thin_border()
                    part_col_letters.append(get_column_letter(col))

                total_col  = col_start + len(ASS_PARTS)
                total_ref  = f"{part_col_letters[0]}{row}:{part_col_letters[-1]}{row}"
                total_cell = ws.cell(row=row, column=total_col)
                total_cell.value       = f"=SUM({total_ref})"
                total_cell.fill        = PatternFill("solid", fgColor="FFFBEB")
                total_cell.font        = Font(color="92400E", bold=True, name="Calibri")
                total_cell.alignment   = _center()
                total_cell.border      = _thin_border()
                total_cell.number_format = "0"
                ass_total_cols.append(get_column_letter(total_col))

            # Assign Avg
            ass_total_ref = "+".join([f"{c}{row}" for c in ass_total_cols])
            aa_cell = ws.cell(row=row, column=ASSIGN_AVG_COL)
            aa_cell.value          = f"=IFERROR(({ass_total_ref})/2,\"\")"
            aa_cell.fill           = PatternFill("solid", fgColor="D1FAE5")
            aa_cell.font           = Font(color="065F46", bold=True, name="Calibri")
            aa_cell.alignment      = _center()
            aa_cell.border         = _thin_border()
            aa_cell.number_format  = "0.0"

        ws.row_dimensions[row].height = 22

    # ── Data Validation: per part column ─────────────────────
    for exp_idx in range(NUM_EXPS):
        col_start = EXP_COLS_START + exp_idx * EXP_STRIDE
        for part_idx, (part_label, part_max) in enumerate(EXP_PARTS):
            col_letter = get_column_letter(col_start + part_idx)
            sqref = f"{col_letter}{DATA_START}:{col_letter}{data_end_row}"
            _add_dv(ws, sqref, part_max)

    if HAS_ASSIGNS:
        for ass_idx in range(2):
            col_start = ASS_COLS_START + ass_idx * ASS_STRIDE
            for part_idx, (part_label, part_max) in enumerate(ASS_PARTS):
                col_letter = get_column_letter(col_start + part_idx)
                sqref = f"{col_letter}{DATA_START}:{col_letter}{data_end_row}"
                _add_dv(ws, sqref, part_max)

    # ── Column widths ─────────────────────────────────────────
    ws.column_dimensions["A"].width = 9   # Roll
    ws.column_dimensions["B"].width = 28  # Name
    for exp_idx in range(NUM_EXPS):
        col_start = EXP_COLS_START + exp_idx * EXP_STRIDE
        for offset in range(len(EXP_PARTS)):
            ws.column_dimensions[get_column_letter(col_start + offset)].width = 7
        ws.column_dimensions[get_column_letter(col_start + len(EXP_PARTS))].width = 8  # total

    ws.column_dimensions[get_column_letter(EXP_AVG_COL)].width = 9

    if HAS_ASSIGNS:
        for ass_idx in range(2):
            col_start = ASS_COLS_START + ass_idx * ASS_STRIDE
            for offset in range(len(ASS_PARTS)):
                ws.column_dimensions[get_column_letter(col_start + offset)].width = 7
            ws.column_dimensions[get_column_letter(col_start + len(ASS_PARTS))].width = 8
        ws.column_dimensions[get_column_letter(ASSIGN_AVG_COL)].width = 9

    # ── Freeze panes ──────────────────────────────────────────
    ws.freeze_panes = f"C{DATA_START}"

    # ── Sheet 2: Instructions ─────────────────────────────────
    wi = wb.create_sheet(title="Instructions")
    wi.sheet_view.showGridLines = False

    wi.merge_cells("A1:C1")
    t = wi["A1"]
    t.value = "📋  How to use this template"
    t.fill  = PRAC_TITLE_FILL
    t.font  = Font(color="FFFFFF", bold=True, size=13, name="Calibri")
    t.alignment = _center()
    wi.row_dimensions[1].height = 32

    instructions = [
        ("1", "Fill the marks for each Experiment part: A (max 3), B (max 4), C (max 4), D (max 4)."),
        ("2", "For Assignments, fill parts A (max 2), B (max 2), C (max 1) for each assignment."),
        ("3", "Do NOT modify Roll No or Student Name columns."),
        ("4", "Do NOT rename column headers — the upload parser depends on them."),
        ("5", "Total and Average columns are auto-calculated — do not type in them."),
        ("6", "Save the file as .xlsx before uploading."),
        ("7", "Upload using the Bulk Upload button on the Marks Management page."),
    ]

    for i, (num, text) in enumerate(instructions, start=2):
        wi.row_dimensions[i].height = 24
        num_cell = wi.cell(row=i, column=1, value=num)
        num_cell.font = Font(bold=True, color="065F46", name="Calibri")
        num_cell.alignment = _center()
        num_cell.fill = PatternFill("solid", fgColor="ECFDF5")
        num_cell.border = _thin_border()

        text_cell = wi.cell(row=i, column=2, value=text)
        text_cell.font = Font(name="Calibri")
        text_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        text_cell.border = _thin_border()
        wi.merge_cells(f"B{i}:C{i}")

    wi.column_dimensions["A"].width = 5
    wi.column_dimensions["B"].width = 80
    wi.column_dimensions["C"].width = 10

    return wb
